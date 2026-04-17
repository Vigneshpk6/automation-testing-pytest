"""
generate_excel.py  —  Creates test_data.xlsx with a 'run' Y/N column on every sheet.
Only rows marked run=Y will be executed by the test suite.
Run: python generate_excel.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()

# ── Style helpers ──────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", start_color="1a1a2e")
HDR_FONT  = Font(bold=True, color="b5f23c", name="Arial", size=11)
BODY_FONT = Font(name="Arial", size=10)
Y_FONT    = Font(name="Arial", size=10, color="27ae60", bold=True)
N_FONT    = Font(name="Arial", size=10, color="e74c3c", bold=True)
OK_FONT   = Font(name="Arial", size=10, color="27ae60", bold=True)
FAIL_FONT = Font(name="Arial", size=10, color="e74c3c", bold=True)
Y_FILL    = PatternFill("solid", start_color="eafaf1")
N_FILL    = PatternFill("solid", start_color="fdf2f2")


def make_header(ws, headers: list[str], widths: list[int]):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[chr(64 + col)].width = w
        c = ws.cell(row=1, column=col, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")


def style_body(ws, result_col: int | None = None, run_col: int | None = None):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = BODY_FONT
        if result_col:
            rc = row[result_col - 1]
            rc.font = OK_FONT if rc.value == "success" else FAIL_FONT
        if run_col:
            rc = row[run_col - 1]
            is_y = str(rc.value).upper() == "Y"
            rc.font  = Y_FONT  if is_y else N_FONT
            rc.fill  = Y_FILL  if is_y else N_FILL
            rc.alignment = Alignment(horizontal="center")


# ════════════════════════════════════════════════════════════════════════════
#  Sheet 1 — LoginData
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "LoginData"
make_header(ws1,
    ["run", "test_case",                   "username", "password",  "expected_result"],
    [6,     38,                             18,         18,          18])

# run=Y → will execute | run=N → skipped
login_data = [
    ("Y", "Valid admin login",              "admin",    "admin123",  "success"),
    ("Y", "Valid user login",               "user",     "pass123",   "success"),
    ("Y", "Wrong password",                 "admin",    "wrongpwd",  "failure"),
    ("Y", "Unknown username",               "nobody",   "admin123",  "failure"),
    ("Y", "Empty credentials",             "",          "",          "failure"),
    ("N", "Missing password only",          "admin",    "",          "failure"),   # ← disabled
    ("N", "Missing username only",          "",         "admin123",  "failure"),   # ← disabled
]
for r in login_data:
    ws1.append(list(r))
style_body(ws1, result_col=5, run_col=1)


# ════════════════════════════════════════════════════════════════════════════
#  Sheet 2 — TaskData
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("TaskData")
make_header(ws2,
    ["run", "test_case",                   "task_name",              "priority", "should_complete"],
    [6,     35,                             30,                        12,         18])

task_data = [
    ("Y", "Add low priority task",          "Buy groceries",          "low",      "Y"),
    ("Y", "Add medium priority task",       "Write test report",      "medium",   "N"),
    ("Y", "Add high + complete it",         "Fix critical bug",       "high",     "Y"),
    ("Y", "Add medium, keep active",        "Code review",            "medium",   "N"),
    ("N", "Add high, keep active",          "Deploy to production",   "high",     "N"),  # disabled
    ("Y", "Add low + complete it",          "Update documentation",   "low",      "Y"),
]
for r in task_data:
    ws2.append(list(r))
style_body(ws2, run_col=1)


# ════════════════════════════════════════════════════════════════════════════
#  Sheet 3 — FilterScenarios
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("FilterScenarios")
make_header(ws3,
    ["run", "filter_name", "expected_visible", "expected_hidden", "notes"],
    [6,     16,             20,                  20,                42])

filter_data = [
    ("Y", "all",       "3", "0", "All 3 tasks visible after seeding"),
    ("Y", "active",    "2", "1", "1 completed → only 2 active shown"),
    ("Y", "completed", "1", "2", "Only the completed task shown"),
]
for r in filter_data:
    ws3.append(list(r))
style_body(ws3, run_col=1)


# ════════════════════════════════════════════════════════════════════════════
#  Sheet 4 — UIChecks
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("UIChecks")
make_header(ws4,
    ["run", "check_name",              "selector",           "expected_text",        "visible"],
    [6,     30,                         24,                    26,                     8])

ui_data = [
    ("Y", "Page title",                "title",               "TaskFlow — Sample App", "Y"),
    ("Y", "Login button text",         "#login-btn",          "Sign In",               "Y"),
    ("Y", "Add button present",        "#add-btn",            "+ Add",                 "Y"),
    ("Y", "All filter button",         "[data-filter=all]",   "All",                   "Y"),
    ("N", "Empty state visible",       "#empty-state",        "",                      "Y"),  # disabled
]
for r in ui_data:
    ws4.append(list(r))
style_body(ws4, run_col=1)


# ════════════════════════════════════════════════════════════════════════════
#  Sheet 5 — DeleteScenarios
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("DeleteScenarios")
make_header(ws5,
    ["run", "test_case",                    "task_name",          "expected_count_after"],
    [6,     38,                              28,                    22])

delete_data = [
    ("Y", "Delete only task → empty state", "Single task",         "0"),
    ("Y", "Delete one of two tasks",        "First of two",        "1"),
    ("N", "Delete all tasks one by one",    "Multi-delete",        "0"),  # disabled
]
for r in delete_data:
    ws5.append(list(r))
style_body(ws5, run_col=1)

wb.save("test_data.xlsx")

print("✅  test_data.xlsx generated:")
for ws in wb.worksheets:
    total = ws.max_row - 1
    enabled = sum(
        1 for row in ws.iter_rows(min_row=2, values_only=True)
        if str(row[0]).upper() == "Y"
    )
    print(f"   • {ws.title:<20} {total} rows  →  {enabled} enabled (Y), {total - enabled} skipped (N)")
